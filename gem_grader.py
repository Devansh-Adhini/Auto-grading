import os
import re
import time
import glob
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook

# --- Configuration ---
GEM_URL = "https://gemini.google.com/gem/6981c34e6a08"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_DIR = os.path.join(BASE_DIR, "tut5/pdf")
PROFILE_DIR = os.path.join(BASE_DIR, "chrome_profile")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "tut5/grades.xlsx")
BATCH_SIZE = 1  # Number of PDFs per batch
DELAY_BETWEEN_BATCHES = 5  # seconds between batches


def setup_driver():
    """Sets up the Chrome driver with the persistent profile."""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    if not os.path.exists(PROFILE_DIR):
        print(f"[ERROR] Chrome profile not found at {PROFILE_DIR}")
        print("Please run create_chrome_profile.py first and log in to Google.")
        return None

    options.add_argument(f"user-data-dir={PROFILE_DIR}")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def open_upload_menu(driver):
    """
    Clicks the '+' (Open upload file menu) button in Gemini's chat input area.
    We know from debugging it has aria-label='Open upload file menu'.
    """
    # Strategy 1: Direct aria-label match (we know this works)
    selectors = [
        '[aria-label="Open upload file menu"]',
        '[aria-label*="upload file menu"]',
        '[aria-label*="Upload file menu"]',
    ]
    for selector in selectors:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, selector)
            btn.click()
            print("  Clicked 'Open upload file menu' button.")
            return True
        except Exception:
            continue

    # Strategy 2: Find first button inside input-container
    try:
        buttons = driver.find_elements(By.XPATH, '//input-container//button')
        if buttons:
            buttons[0].click()
            print("  Clicked first button in input-container.")
            return True
    except Exception:
        pass

    # Strategy 3: JavaScript fallback
    try:
        result = driver.execute_script("""
            const buttons = document.querySelectorAll('button');
            for (const btn of buttons) {
                const label = (btn.getAttribute('aria-label') || '').toLowerCase();
                if (label.includes('upload file menu') || label.includes('open upload')) {
                    btn.click();
                    return 'clicked';
                }
            }
            return null;
        """)
        if result:
            print("  Clicked upload menu button via JavaScript.")
            return True
    except Exception:
        pass

    print("  [ERROR] Could not find the upload menu button.")
    return False


def click_upload_files_option(driver):
    """
    Clicks the 'Upload files' option from the attachment popup menu.
    """
    time.sleep(1.5)  # Wait for menu animation

    # Try various selectors for the "Upload files" menu item
    xpaths = [
        '//*[contains(text(), "Upload files")]',
        '//*[contains(text(), "Upload file")]',
        '//span[contains(text(), "Upload")]',
        '//*[@role="menuitem"][contains(., "Upload")]',
    ]

    for xpath in xpaths:
        try:
            elements = driver.find_elements(By.XPATH, xpath)
            for el in elements:
                if el.is_displayed():
                    el.click()
                    print("  Clicked 'Upload files' menu item.")
                    return True
        except Exception:
            continue

    # JavaScript fallback
    try:
        result = driver.execute_script("""
            const allElements = document.querySelectorAll('*');
            for (const el of allElements) {
                const text = el.textContent.trim();
                if ((text === 'Upload files' || text === 'Upload file') && el.children.length <= 2) {
                    el.click();
                    return 'clicked';
                }
            }
            return null;
        """)
        if result:
            print("  Clicked 'Upload files' via JavaScript.")
            return True
    except Exception:
        pass

    print("  [ERROR] Could not find 'Upload files' menu item.")
    return False


def handle_file_dialog(file_paths):
    """
    Uses pyautogui to interact with the native Windows file dialog.
    Types the file path(s) into the 'File name' field and presses Enter.
    """
    import pyautogui
    import pyperclip

    # Wait for the OS file dialog to fully open
    time.sleep(3)

    if len(file_paths) == 1:
        # Single file: type the full absolute path
        full_path = os.path.abspath(file_paths[0])
        print(f"  Typing file path into dialog: {os.path.basename(full_path)}")
        pyperclip.copy(full_path)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.5)
        pyautogui.press('enter')
    else:
        # Multiple files: type directory path first, then select files
        directory = os.path.dirname(os.path.abspath(file_paths[0]))
        filenames = [os.path.basename(os.path.abspath(p)) for p in file_paths]

        # First, navigate to the directory
        print(f"  Navigating to directory: {directory}")
        pyperclip.copy(directory)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.5)
        pyautogui.press('enter')
        time.sleep(2)  # Wait for directory to load

        # Now type the filenames with quotes for multi-select
        file_str = ' '.join(f'"{f}"' for f in filenames)
        print(f"  Selecting files: {file_str}")
        pyperclip.copy(file_str)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.5)
        pyautogui.press('enter')

    print("  File dialog handled.")
    time.sleep(2)


def upload_pdfs(driver, pdf_paths):
    """
    Uploads PDF files to the Gemini gem chat.
    Flow: Click '+' menu -> Click 'Upload files' -> Handle OS file dialog with pyautogui.
    """
    print(f"  Uploading {len(pdf_paths)} PDF(s)...")

    # Step 1: Click the '+' button to open attachment menu
    if not open_upload_menu(driver):
        return False

    # Step 2: Click 'Upload files' from the dropdown menu
    if not click_upload_files_option(driver):
        return False

    # Step 3: Handle the native OS file dialog with pyautogui
    print("  Waiting for OS file dialog to open...")
    handle_file_dialog(pdf_paths)

    # Step 4: Wait for files to be processed/attached in Gemini
    print("  Waiting for files to attach in Gemini...")
    time.sleep(8)

    print("  Files uploaded successfully.")
    return True


def send_message(driver):
    """
    Sends the message (triggers the gem to process the uploaded files).
    The gem already has its prompt, so we just need to press Enter or click Send.
    """
    print("  Sending message to Gemini gem...")

    # Try to find the text input area using the provided XPath
    text_input_xpath = "/html/body/chat-app/main/side-navigation-v2/bard-sidenav-container/bard-sidenav-content/div[2]/div/div[2]/bots-chat-window/chat-window/div/input-container/fieldset/input-area-v2/div/div/div[1]/div/div/rich-textarea/div[1]/p"

    try:
        text_input = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, text_input_xpath))
        )
        text_input.click()
        time.sleep(0.5)
        # Type a simple trigger message and press Enter
        text_input.send_keys("Grade these PDFs")
        time.sleep(1)
        text_input.send_keys(Keys.RETURN)
        print("  Message sent via Enter key.")
        return True
    except Exception as e1:
        print(f"  XPath method failed: {e1}")
        # Fallback: try to find send button
        try:
            send_btn = driver.find_element(By.CSS_SELECTOR, '[aria-label="Send message"]')
            send_btn.click()
            print("  Message sent via send button.")
            return True
        except Exception as e2:
            print(f"  Send button fallback failed: {e2}")
            # Another fallback: try contenteditable div
            try:
                rich_textarea = driver.find_element(By.CSS_SELECTOR, '.ql-editor, [contenteditable="true"], rich-textarea')
                rich_textarea.click()
                time.sleep(0.5)
                ActionChains(driver).send_keys("Grade these PDFs").perform()
                time.sleep(0.5)
                ActionChains(driver).send_keys(Keys.RETURN).perform()
                print("  Message sent via contenteditable fallback.")
                return True
            except Exception as e3:
                print(f"  [ERROR] All send methods failed: {e3}")
                return False


def is_still_generating(driver):
    """
    Checks if Gemini is still generating a response by looking for
    the stop/generating button or animation indicators.
    """
    try:
        # Look for stop button (blue square that appears during generation)
        stop_indicators = [
            '[aria-label="Stop"]',
            '[aria-label="stop"]',
            '[aria-label*="Stop generating"]',
            '.loading-indicator',
            '.generating',
        ]
        for selector in stop_indicators:
            try:
                el = driver.find_element(By.CSS_SELECTOR, selector)
                if el.is_displayed():
                    return True
            except:
                continue

        # Also check via JS for the send button state (it becomes a stop button during gen)
        result = driver.execute_script("""
            // If there's a stop button visible, generation is ongoing
            const buttons = document.querySelectorAll('button');
            for (const btn of buttons) {
                const label = (btn.getAttribute('aria-label') || '').toLowerCase();
                if (label.includes('stop') && btn.offsetParent !== null) {
                    return true;
                }
            }
            return false;
        """)
        return result == True
    except:
        return False


def wait_for_response(driver, timeout=600):
    """
    Waits for Gemini to finish generating its response.
    - Waits 30s initially for generation to begin
    - Checks for the stop/generating button to know if still generating
    - Monitors text stability (no change for ~15 seconds)
    - Ignores gem header text (short static content)
    Returns the response text.
    """
    print("  Waiting for Gemini response (this may take a while)...")

    # Ignore these patterns - they're gem header text, not actual responses
    IGNORE_PATTERNS = [
        'grader',
        'custom gem',
        'analysis',
        'grader said',
    ]

    def is_header_text(text):
        """Returns True if the text looks like gem header, not actual response."""
        if not text:
            return True
        # If text is very short and matches header patterns, ignore it
        if len(text) < 50:
            text_lower = text.lower().strip()
            lines = [l.strip().lower() for l in text_lower.split('\n') if l.strip()]
            # Check if all lines are header text
            if all(any(pat in line for pat in IGNORE_PATTERNS) or len(line) < 5 for line in lines):
                return True
        return False

    # Wait 30 seconds before even starting to check
    # Gemini needs time to process PDFs
    print("  Giving Gemini 30 seconds to start processing PDFs...")
    time.sleep(30)

    last_text = ""
    stable_count = 0
    stable_threshold = 5  # 5 consecutive checks with same text (5 * 3s = 15s)
    max_attempts = timeout // 3  # Check every 3 seconds

    for attempt in range(max_attempts):
        try:
            # Check if still generating
            generating = is_still_generating(driver)
            if generating and attempt % 5 == 0:
                print(f"  [Check {attempt}] Gemini is still generating...")

            # Try multiple selectors for the response area
            response_elements = []
            selectors = [
                'message-content',
                'model-response',
                '.response-container',
                '.markdown-main-panel',
                '[class*="response"]',
                '[class*="message-text"]',
            ]

            for selector in selectors:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        response_elements = elements
                        break
                except:
                    continue

            if not response_elements:
                try:
                    response_elements = driver.find_elements(
                        By.XPATH, '//message-content | //div[contains(@class, "response")] | //div[contains(@class, "markdown")]'
                    )
                except:
                    pass

            if response_elements:
                # Get text from the last response element
                current_text = response_elements[-1].text.strip()

                # Skip if it's just header text
                if is_header_text(current_text):
                    if attempt % 10 == 0:
                        print(f"  [Check {attempt}] Only header text so far, waiting...")
                    time.sleep(3)
                    continue

                if current_text and current_text == last_text and not generating:
                    stable_count += 1
                    if stable_count >= stable_threshold:
                        print(f"  Response stabilized after ~{30 + attempt * 3} seconds.")
                        return current_text
                else:
                    stable_count = 0
                    last_text = current_text

                if attempt % 5 == 0 and current_text:
                    preview = current_text[:100] + "..." if len(current_text) > 100 else current_text
                    print(f"  [Check {attempt}] Response ({len(current_text)} chars): {preview}")

        except Exception as e:
            if attempt % 10 == 0:
                print(f"  [Check {attempt}] Waiting... ({e})")

        time.sleep(3)

    print("  [WARNING] Response did not stabilize within timeout. Returning last captured text.")
    return last_text


def parse_response(response_text, pdf_names):
    """
    Parses the Gemini response into structured data.
    Expected CSV format (no commas inside fields):
    name,roll_no,pdf_name,grade,reason_for_deduction,missing_questions
    Example: alex,202311026,math_pdf,8.5,wrong calculation in Q3,Q11 Q12
    """
    results = []

    if not response_text:
        print("  [WARNING] Empty response received.")
        return results

    print("\n  --- Raw Response ---")
    print(f"  {response_text}")
    print("  --- End Response ---\n")

    lines = response_text.strip().split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Split by comma — exactly 6 fields expected
        parts = [p.strip() for p in line.split(',')]

        if len(parts) >= 4:
            result = {
                'Name': parts[0],
                'Roll No': parts[1] if len(parts) > 1 else '',
                'PDF Name': parts[2] if len(parts) > 2 else '',
                'Grade': parts[3] if len(parts) > 3 else '',
                'Reason for Deduction': parts[4] if len(parts) > 4 else 'None',
                'Missing Questions': parts[5] if len(parts) > 5 else 'None',
            }
            results.append(result)
            print(f"  Parsed: {result['Name']} | {result['Roll No']} | Grade: {result['Grade']}")

    if not results:
        print("  [WARNING] Could not parse structured data. Saving raw response.")
        results.append({
            'Name': 'PARSE_ERROR',
            'Roll No': '',
            'PDF Name': ', '.join(pdf_names),
            'Grade': '',
            'Reason for Deduction': '',
            'Missing Questions': '',
            'Raw Response': response_text
        })

    return results


def save_to_excel(results, output_path):
    """
    Saves parsed results to an Excel file.
    Creates the file if it doesn't exist, appends if it does.
    """
    headers = ['Name', 'Roll No', 'PDF Name', 'Grade', 'Reason for Deduction', 'Missing Questions']

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
        print(f"  Appending to existing Excel file: {output_path}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(headers)

        # Style headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        print(f"  Created new Excel file: {output_path}")

    for result in results:
        row = [result.get(h, '') for h in headers]
        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"  [SUCCESS] Saved {len(results)} record(s) to {output_path}")


def start_new_chat(driver):
    """
    Navigates to the gem URL to start a fresh chat.
    """
    print("  Starting new chat with gem...")
    driver.get(GEM_URL)
    time.sleep(8)  # Wait for page to fully load


def main():
    # Validate PDF directory
    if not os.path.exists(PDF_DIR):
        os.makedirs(PDF_DIR)
        print(f"[INFO] Created PDF directory: {PDF_DIR}")
        print("[INFO] Please place your student PDFs in this folder and run again.")
        return

    # Collect all PDFs
    pdf_files = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
    if not pdf_files:
        print(f"[ERROR] No PDF files found in {PDF_DIR}")
        print("Please place student PDFs in the 'pdf' folder and run again.")
        return

    print(f"Found {len(pdf_files)} PDF(s) total.")

    # --- Resume: skip already-graded PDFs ---
    already_graded = set()
    if os.path.exists(OUTPUT_EXCEL):
        try:
            wb = load_workbook(OUTPUT_EXCEL)
            ws = wb.active
            # Read the 'PDF Name' column (column 3) skipping header
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
                if row[0]:
                    already_graded.add(row[0].strip())
            wb.close()
            if already_graded:
                print(f"Resuming: {len(already_graded)} PDF(s) already graded, skipping them.")
        except Exception as e:
            print(f"[WARNING] Could not read existing grades file: {e}")

    # Filter out already-graded PDFs
    pdf_files = [p for p in pdf_files if os.path.basename(p) not in already_graded]

    if not pdf_files:
        print("All PDFs have already been graded! Nothing to do.")
        return

    print(f"{len(pdf_files)} PDF(s) remaining to grade:")
    for i, f in enumerate(pdf_files):
        print(f"  {i+1}. {os.path.basename(f)}")

    # Create batches
    batches = []
    for i in range(0, len(pdf_files), BATCH_SIZE):
        batches.append(pdf_files[i:i + BATCH_SIZE])
    print(f"\nWill process {len(batches)} batch(es) of up to {BATCH_SIZE} PDFs each.\n")

    # Setup driver
    try:
        driver = setup_driver()
        if not driver:
            return
    except Exception as e:
        print(f"\n[ERROR] Failed to start Chrome: {e}")
        print("Make sure all other Chrome instances with this profile are closed.")
        return

    try:
        all_results = []

        for batch_idx, batch in enumerate(batches):
            pdf_names = [os.path.basename(p) for p in batch]
            print(f"\n{'='*60}")
            print(f"BATCH {batch_idx + 1}/{len(batches)}: {', '.join(pdf_names)}")
            print(f"{'='*60}")

            # Navigate to gem only on first batch
            if batch_idx == 0:
                start_new_chat(driver)
                print("\n" + "="*60)
                print("ACTION REQUIRED: Make sure the Gemini gem page has loaded.")
                print("If there are any popups or consent dialogs, please dismiss them.")
                input("Press ENTER in this terminal when ready to start grading...")
                print("="*60 + "\n")

            # Upload PDFs
            success = upload_pdfs(driver, batch)
            if not success:
                print(f"  [ERROR] Failed to upload PDFs for batch {batch_idx + 1}. Skipping.")
                continue

            # Send message to trigger grading
            success = send_message(driver)
            if not success:
                print(f"  [ERROR] Failed to send message for batch {batch_idx + 1}. Skipping.")
                continue

            # Wait for and capture response
            response = wait_for_response(driver)

            if response:
                # Parse the response
                results = parse_response(response, pdf_names)
                all_results.extend(results)

                # Save after each batch (incremental saving)
                save_to_excel(results, OUTPUT_EXCEL)
            else:
                print(f"  [ERROR] No response received for batch {batch_idx + 1}.")

            # Delay between batches
            if batch_idx < len(batches) - 1:
                print(f"\n  Waiting {DELAY_BETWEEN_BATCHES} seconds before next batch...")
                time.sleep(DELAY_BETWEEN_BATCHES)

        print(f"\n{'='*60}")
        print(f"GRADING COMPLETE")
        print(f"Total records processed: {len(all_results)}")
        print(f"Results saved to: {OUTPUT_EXCEL}")
        print(f"{'='*60}")

    except Exception as e:
        print(f"\n[GLOBAL ERROR] {e}")
        import traceback
        traceback.print_exc()

    finally:
        print("\nPress Enter to close the browser...")
        input()
        driver.quit()


if __name__ == "__main__":
    main()
